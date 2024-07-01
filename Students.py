import datetime

from data import wb

now = datetime.datetime.now()
import matplotlib.pyplot as mplt


class Student:
    """פעולה בונה לתלמידה"""
    def __init__(self,name,classname,date,time,lateNum,oldLate):
        self.name=name
        self.late = {
            "classname":classname,
            "date": date,
            "time": time,
            "lateNum": lateNum,
            "oldLate":oldLate
        }
        print("name",self.name,"late",self.late,"oldLate")

"""מילון תלמידות"""
class StudentsArr:
    def __init__(self):
        self.dictstudent = [
            Student(name="rut",classname="a",date="0.0.0",time="0:0",lateNum=5,oldLate=0),
            Student(name="shoshana", classname="b",date="0.0.0",time="0:0",lateNum=0,oldLate=0),
            Student(name="rachel", classname="b",date="0.0.0",time="0:0",lateNum=9,oldLate=0),
            Student(name="bella", classname="a",date="0.0.0",time="0:0",lateNum=0,oldLate=0),
            Student(name="naama", classname="c",date="0.0.0",time="0:0",lateNum=0,oldLate=0),
            Student(name="meir", classname="a", date="0.0.0", time="0:0", lateNum=3,oldLate=0)

        ]

    def diagram(self):
        dict = []
        for i in self.dictstudent:
            late = i.late["lateNum"]
            dict.append(late)
        return dict

    def namess(self):
        print("i am here")
        dict = []
        for i in self.dictstudent:
            name = i.name
            dict.append(name)
            # print(dict)
        return dict
    # def drawmplt(self):
    #     d=[]
    #     for i in self.dictstudent:
    #        print(i.lateNum())

        #print(d)#
        # mplt.plot()

    def add_late_for_student(self,name):
        """פונקציה המעדכנת את האיחור"""
        date = now.strftime("%Y-%m-%d")
        time = now.strftime(" %H:%M:%S")
        for i in self.dictstudent:
            if name==i.name:
                i.late["date"] =date
                i.late["time"] =time
                i.late["lateNum"]=i.late["lateNum"]+1
            if i.late["lateNum"]==10:
                i.oldLate=i.oldLate=1

    def num_of_lates(self,name):
        """פונקציה הבודקת האם צריך לקנות כרטיסיה """
        for i in self.dictstudent:
            if name == i.name:
                if i.late["lateNum"]==10:
                    i.late["lateNum"]=0
                    return True
        return False

    def get_students_names(self):
        """מחזירה את מערך התלמידות ע"י פונקציית למדה"""
        printinlamda = list(map(lambda x: x.name, self.dictstudent))
        print(printinlamda)

    def get_students_names_in_class(self,classNum):
        """מחזירה את מערך התלמידות"""
        for i in self.dictstudent:
            if i.late["classname"]==classNum:
                print(i.name)

    def min_late(self,):
        """פונקציה המחזירה את שמות הבנות שלא אחרו"""
        print([item.name for item in self.dictstudent if item.late["lateNum"]==0])

    def find_student(self,name):
        """בודק האם השם שנכנס קיים במערך התלמידות"""
        for i in self.dictstudent:
                       if name==i.name:
                           print()
                           return True
                       else:
                        continue
        return False

    def maxlateclass(self):
        """מחזיר את הכיתה עם הכי הרבה איחורים"""
        max=0
        ii=0
        dictclass={"a":0,"b":0,"c":0}
        for i in self.dictstudent:
            p=i.late["classname"]
            dictclass[p]=dictclass[p]+i.late["lateNum"]
        for i in dictclass:
            if dictclass[i]>max:
                ii=i
                max=dictclass[i]
        print(f"class:{ii}  max:{max}")

    def return_my_lates(self,name):
        for i in self.dictstudent:
            if name==i.name:
                late=i.late["lateNum"]
                late=late+1
                print(late)
                return late

    def Expention(self,):
        """פונקציית שגיאה"""
        try:
            num=int(input("please enter a number"))
            for i in self.dictstudent:
                if num == i.name:
                    i.name = num
        except Exception as e:
            print(f"the exception was {e}")
        else:
            print("no error happened")
        finally:
            print(f"we tried dividing 12 by something")

    def return_dict(self):
        return self.dictstudent

    import openpyxl as xl
    from openpyxl.cell import read_only
    from openpyxl.reader.excel import ExcelReader

    file_name = "studentss.xlsx"
    wb = xl.load_workbook(r"studentss.xlsx")
    reader = ExcelReader(file_name, read_only)


    def output_data_excel(self):
        import openpyxl as xl

        wb = xl.load_workbook(r"studentss.xlsx")
        sheet2 = wb['גיליון1']
        a = 1

        for i in self.dictstudent:
            sheet2.cell(a, 1).value =i.name
            a = a + 1
        a = 1
        for i in self.dictstudent:
            sheet2.cell(a, 2).value = i.late["lateNum"]
            a = a + 1
        wb.save(r"studentss.xlsx")


