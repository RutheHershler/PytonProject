
import openpyxl as xl
from openpyxl.cell import read_only
from openpyxl.reader.excel import ExcelReader

file_name="studentss.xlsx"
wb=xl.load_workbook(r"studentss.xlsx")
reader = ExcelReader(file_name, read_only)
class excel:
    print("smjnj datatatata")
    def output_data_excel(dictstudent):
        sheet2=wb['גיליון1']
        a=1

        for i in dictstudent:
            sheet2.cell(a, 0).value=i.name
            a=a+1
        a=2
        for i in dictstudent:
            sheet2.cell(a, 0).value=i.late["lateNum"]
            a=a+1
        wb.save(r"studentss.xlsx")


